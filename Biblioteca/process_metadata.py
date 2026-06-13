import os
import re
import json
import time
import zipfile
import urllib.request
import urllib.parse
import xml.etree.ElementTree as ET
import openpyxl
import requests
from pypdf import PdfReader

# Configuración de rutas
base_dir = os.path.dirname(os.path.abspath(__file__))
excel_input = os.path.join(base_dir, "GELATINA - LIBROS POR LIMPIAR.xlsx")
excel_output = os.path.join(base_dir, "GELATINA - LIBROS PROCESADOS.xlsx")
books_dir = os.path.join(base_dir, "Libros")
cache_file = os.path.join(base_dir, "metadata_cache.json")
env_file = os.path.join(base_dir, ".env")

# Mapeo de categorías a géneros homologados del Excel
GENRE_MAPPING = {
    # Erótico / Romance
    "erotic": "Romance erótico",
    "erotica": "Romance erótico",
    "romance": "Romance contemporáneo",
    "love": "Romance contemporáneo",
    "regency": "Romance histórico",
    "historical romance": "Romance histórico",
    "mafia": "Romance / mafia",
    # Policial / Misterio
    "mystery": "Policial / misterio",
    "detective": "Policial / misterio",
    "crime": "Policial / misterio",
    "police": "Policial / misterio",
    "thriller": "Thriller / suspense",
    "suspense": "Thriller / suspense",
    "novela negra": "Western / novela negra",
    "western": "Western / novela negra",
    # Fantasía / Ciencia Ficción
    "fantasy": "Fantasía",
    "urban fantasy": "Fantasía urbana juvenil",
    "paranormal": "Fantasía paranormal",
    "science fiction": "Ciencia ficción",
    "sci-fi": "Ciencia ficción",
    "dystopian": "Distopía / ciencia ficción juvenil",
    "dystopia": "Distopía / ciencia ficción juvenil",
    # Negocios / Finanzas
    "business": "Emprendimiento",
    "economics": "Administración / negocios",
    "finance": "Educación financiera",
    "self-help": "Autoayuda / desarrollo personal",
    "personal development": "Autoayuda / desarrollo personal",
    "success": "Autoayuda / Éxito personal",
    "leadership": "Desarrollo personal / liderazgo",
    # Gastronomía
    "cooking": "Gastronomía / Repostería",
    "cookery": "Gastronomía / Repostería",
    "baking": "Gastronomía / Panadería",
    "recipes": "Cocina / recetario",
    # Otros
    "history": "Novela histórica",
    "biography": "Biografía / negocios",
    "psychology": "Psicología / Desarrollo personal",
    "education": "Académico / educación",
    "meditation": "Mindfulness / meditación",
    "yoga": "Yoga / espiritualidad",
    "spirituality": "Espiritualidad",
    "esoterism": "Esoterismo / Filosofía oculta",
    "philosophy": "Ensayo / Filosofía",
    "juvenile": "Fantasía juvenil",
    "young adult": "Fantasía juvenil / aventura",
    "humor": "Humor"
}

# Variable de control de fallos consecutivos en Google Books
consecutive_google_failures = 0
google_api_enabled = True

def load_env():
    """Carga variables de entorno desde un archivo .env si existe."""
    env = {}
    if os.path.exists(env_file):
        try:
            with open(env_file, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#"):
                        parts = line.split("=", 1)
                        if len(parts) == 2:
                            env[parts[0].strip()] = parts[1].strip()
        except Exception as e:
            print(f"[ENV] Error al leer archivo .env: {e}")
    return env

def load_cache():
    """Carga la caché persistente desde el disco."""
    if os.path.exists(cache_file):
        try:
            with open(cache_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"[CACHE] Error al cargar la caché: {e}")
    return {}

def save_cache(cache):
    """Guarda la caché persistente en el disco."""
    try:
        with open(cache_file, "w", encoding="utf-8") as f:
            json.dump(cache, f, indent=4, ensure_ascii=False)
    except Exception as e:
        print(f"[CACHE] Error al guardar la caché: {e}")

def clean_string(s):
    """Normaliza texto para comparaciones y llaves de caché."""
    if not s:
        return ""
    s = s.lower().strip()
    replacements = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ü": "u", "ñ": "n"
    }
    for char, replacement in replacements.items():
        s = s.replace(char, replacement)
    s = re.sub(r'[^a-z0-9\s]', '', s)
    return re.sub(r'\s+', ' ', s).strip()

def clean_for_query(text):
    """Limpia caracteres extraños y emojis antes de consultar APIs."""
    if not text:
        return ""
    # Quitar paréntesis con números al final (ej. "Walter Riso (6)" -> "Walter Riso")
    text = re.sub(r'\s*\(\d+\)\s*$', '', text)
    text = re.sub(r'[^\w\s\.,;:\-\?!¿¡]', '', text)
    return text.strip()

def parse_filename(filename):
    """Extrae un título y autor sugeridos a partir del nombre del archivo físico."""
    # Remover la extensión real final
    name, ext = os.path.splitext(filename)
    
    # Limpiar extensiones dobles y sufijos como " · versión 1" o "-1" o "_filename"
    name = re.sub(r'\.pdf|\.epub|\.mobi', '', name, flags=re.IGNORECASE)
    name = re.sub(r' · versi[oó]n \d+', '', name, flags=re.IGNORECASE)
    name = re.sub(r'-\d+(-\d+)*$', '', name)
    name = re.sub(r'_filename= UTF-8_', '', name, flags=re.IGNORECASE)
    name = re.sub(r'_filename_= UTF-8_[A-Fa-f0-9_]*', '', name, flags=re.IGNORECASE)
    
    # Quitar guiones bajos de los bordes
    name = name.strip('_').strip()
    
    # Limpiar prefijos comunes como (Fjallbacka 01), 01-, etc.
    name = re.sub(r'^[\(\[\{][^\)\]\}]+[\)\]\}]\s*', '', name)
    name = re.sub(r'^\d+[\.\-\s_]*', '', name)
    name = re.sub(r'^[^\w\s]+\s*', '', name)
    
    # Quitar guiones bajos resultantes y espacios de nuevo
    name = name.strip('_').strip()
    
    # Intentar separar por delimitadores estándar de guiones
    parts = re.split(r'\s+-\s+|\s*—\s*|\s+–\s+', name)
    if len(parts) >= 2:
        title = parts[0].strip('_').strip()
        author = parts[1].strip('_').strip()
        return title, author
    
    # Si tiene guiones bajos, intentar separar por ellos
    if "_" in name and " " not in name:
        parts_underscore = name.split("_")
        if len(parts_underscore) >= 2:
            return " ".join(parts_underscore[:-1]), parts_underscore[-1]
            
    return name, "Desconocido"

def get_epub_metadata(file_path):
    """Extrae título, autor, tema (género) y descripción (sinopsis) de un EPUB usando XML."""
    try:
        with zipfile.ZipFile(file_path, 'r') as epub:
            container_xml = epub.read('META-INF/container.xml')
            root = ET.fromstring(container_xml)
            ns = {'ns': 'urn:oasis:names:tc:opendocument:xmlns:container'}
            rootfile = root.find('.//ns:rootfile', ns)
            if rootfile is None:
                return None, None, None, None
            
            opf_path = rootfile.attrib['full-path']
            opf_data = epub.read(opf_path)
            opf_root = ET.fromstring(opf_data)
            
            ns_opf = {
                'opf': 'http://www.idpf.org/2007/opf',
                'dc': 'http://purl.org/dc/elements/1.1/'
            }
            
            title_el = opf_root.find('.//dc:title', ns_opf)
            creator_el = opf_root.find('.//dc:creator', ns_opf)
            subject_el = opf_root.find('.//dc:subject', ns_opf)
            desc_el = opf_root.find('.//dc:description', ns_opf)
            
            title = title_el.text.strip() if title_el is not None and title_el.text else None
            author = creator_el.text.strip() if creator_el is not None and creator_el.text else None
            subject = subject_el.text.strip() if subject_el is not None and subject_el.text else None
            desc = desc_el.text.strip() if desc_el is not None and desc_el.text else None
            
            return title, author, subject, desc
    except Exception:
        return None, None, None, None

def get_pdf_metadata(file_path):
    """Extrae metadatos locales de un PDF, interpretando /Subject largo como descripción."""
    try:
        reader = PdfReader(file_path)
        meta = reader.metadata
        title = meta.title.strip() if meta and meta.title else None
        author = meta.author.strip() if meta and meta.author else None
        subject = meta.get('/Subject', None) if meta else None
        desc = meta.get('/Description', None) if meta else None
        
        if isinstance(subject, bytes):
            subject = subject.decode('utf-8', errors='ignore')
        if subject:
            subject = subject.strip()
            
        if isinstance(desc, bytes):
            desc = desc.decode('utf-8', errors='ignore')
        if desc:
            desc = desc.strip()
            
        # Si subject es largo, usualmente es una descripción en PDFs mal formateados
        if subject and len(subject) > 50 and not desc:
            desc = subject
            subject = None
            
        if title and (title.lower().startswith("untitled") or title.lower().startswith("microsoft word") or len(title) < 2):
            title = None
        if author and (len(author) < 2 or "scanner" in author.lower() or "adobe" in author.lower()):
            author = None
            
        return title, author, subject, desc
    except Exception:
        return None, None, None, None

def map_raw_genre(raw_category):
    """Mapea una categoría cruda (en inglés o español) al género del Excel."""
    if not raw_category:
        return "No especificado"
    cat = raw_category.lower()
    for key, val in GENRE_MAPPING.items():
        if key in cat:
            return val
    return raw_category.strip().capitalize()

def query_google_books(title, author, api_key=None):
    """Consulta la API de Google Books con soporte de API Key y reintentos (429)."""
    global consecutive_google_failures, google_api_enabled
    if not google_api_enabled:
        return None, None
        
    q_title = clean_for_query(title)
    q_author = clean_for_query(author)
    if not q_title or q_title == "Desconocido":
        return None, None
        
    query = f"intitle:{q_title}"
    if q_author and q_author != "Desconocido":
        query += f"+inauthor:{q_author}"
        
    url = "https://www.googleapis.com/books/v1/volumes"
    params = {
        "q": query,
        "maxResults": 1
    }
    if api_key:
        params["key"] = api_key
        
    headers = {'User-Agent': 'ImperiusDraconisBot/1.0 (contact@example.com)'}
    
    max_retries = 3
    retry_delay = 4  # segundos
    
    for attempt in range(max_retries):
        try:
            r = requests.get(url, params=params, headers=headers, timeout=8)
            if r.status_code == 200:
                data = r.json()
                consecutive_google_failures = 0  # Éxito: resetear fallos
                if "items" in data and len(data["items"]) > 0:
                    volume_info = data["items"][0]["volumeInfo"]
                    categories = volume_info.get("categories", [])
                    description = volume_info.get("description", "")
                    
                    suggested_genre = "No especificado"
                    if categories:
                        suggested_genre = map_raw_genre(categories[0])
                    return suggested_genre, description
                else:
                    return None, None
            elif r.status_code == 429:
                resp_text = r.text
                if "quota" in resp_text.lower() or "limit" in resp_text.lower() or attempt == max_retries - 1:
                    print("  [Google Books] Cuota o límite excedido. Desactivando Google Books temporalmente.")
                    google_api_enabled = False
                    break
                print(f"  [Google Books 429] Límite superado. Reintentando ({attempt+1}/{max_retries}) en {retry_delay}s...")
                time.sleep(retry_delay)
                retry_delay *= 2
            else:
                print(f"  [Google Books HTTP {r.status_code}] para '{title}'")
                break
        except Exception as e:
            consecutive_google_failures += 1
            print(f"  [Google Books Exception] {e}")
            break
            
    consecutive_google_failures += 1
    if consecutive_google_failures >= 3:
        print("  [SYSTEM] Desactivando consultas directas a Google Books por tasa de fallos.")
        google_api_enabled = False
        
    return None, None

def query_wikidata_wikipedia(title, author):
    """Consulta Wikidata para obtener géneros en español y Wikipedia en español para sinopsis."""
    q_title = clean_for_query(title)
    q_author = clean_for_query(author)
    if not q_title or q_title == "Desconocido":
        return None, None
        
    search_query = f"{q_title} {q_author}" if q_author and q_author != "Desconocido" else q_title
    headers = {'User-Agent': 'ImperiusDraconisBot/1.0 (contact@example.com)'}
    url = "https://www.wikidata.org/w/api.php"
    
    params = {
        "action": "wbsearchentities",
        "search": search_query,
        "language": "es",
        "format": "json",
        "type": "item"
    }
    
    try:
        r = requests.get(url, params=params, headers=headers, timeout=8)
        if r.status_code != 200:
            return None, None
            
        search_results = r.json().get("search", [])
        if not search_results and q_author and q_author != "Desconocido":
            params["search"] = q_title
            r = requests.get(url, params=params, headers=headers, timeout=8)
            if r.status_code == 200:
                search_results = r.json().get("search", [])
                
        if not search_results:
            return None, None
            
        qid = search_results[0]["id"]
        detail_params = {
            "action": "wbgetentities",
            "ids": qid,
            "languages": "es|en",
            "props": "labels|claims|sitelinks",
            "format": "json"
        }
        
        rd = requests.get(url, params=detail_params, headers=headers, timeout=8)
        if rd.status_code != 200:
            return None, None
            
        entity = rd.json().get("entities", {}).get(qid, {})
        
        claims = entity.get("claims", {})
        genre_qids = []
        p136 = claims.get("P136", [])
        for claim in p136:
            val = claim.get("mainsnak", {}).get("datavalue", {}).get("value", {}).get("id")
            if val:
                genre_qids.append(val)
                
        genre_label = "No especificado"
        if genre_qids:
            genre_params = {
                "action": "wbgetentities",
                "ids": "|".join(genre_qids[:3]),
                "props": "labels",
                "languages": "es|en",
                "format": "json"
            }
            rg = requests.get(url, params=genre_params, headers=headers, timeout=8)
            if rg.status_code == 200:
                g_data = rg.json()
                genre_labels = []
                for g_qid in genre_qids[:3]:
                    g_ent = g_data.get("entities", {}).get(g_qid, {})
                    g_lbl = g_ent.get("labels", {}).get("es", {}).get("value") or g_ent.get("labels", {}).get("en", {}).get("value")
                    if g_lbl:
                        genre_labels.append(g_lbl)
                if genre_labels:
                    genre_label = ", ".join(genre_labels)
                    
        sitelinks = entity.get("sitelinks", {})
        eswiki = sitelinks.get("eswiki", {})
        wiki_title = eswiki.get("title")
        
        synopsis = ""
        if wiki_title:
            wiki_url = "https://es.wikipedia.org/w/api.php"
            wiki_params = {
                "action": "query",
                "prop": "extracts",
                "exintro": True,
                "explaintext": True,
                "titles": wiki_title,
                "format": "json"
            }
            rw = requests.get(wiki_url, params=wiki_params, headers=headers, timeout=8)
            if rw.status_code == 200:
                pages = rw.json().get("query", {}).get("pages", {})
                for page_id, page in pages.items():
                    synopsis = page.get("extract", "")
                    break
                    
        return map_raw_genre(genre_label), synopsis
        
    except Exception as e:
        print(f"  [Wikidata Exception] {e}")
        
    return None, None

def query_open_library(title, author):
    """Bypassed for speed since Open Library lacks Spanish book summaries and times out."""
    return None, None

def load_excel_data(file_path):
    """Carga los libros ya revisados del archivo de Excel."""
    reviewed_books = []
    if not os.path.exists(file_path):
        print(f"El archivo {file_path} no existe.")
        return reviewed_books
        
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        reviewed_books.append({
            "item": row[0],
            "titulo": row[1],
            "autor": row[2] if row[2] else "Desconocido",
            "formato": row[3],
            "genero": row[4] if row[4] else "No especificado"
        })
    return reviewed_books

def scan_physical_files(directory):
    """Escanea recursivamente los archivos físicos del directorio Libros."""
    files_list = []
    if not os.path.exists(directory):
        print(f"La carpeta de libros físicos no existe: {directory}")
        return files_list
        
    for root, dirs, files in os.walk(directory):
        for file in files:
            name, ext = os.path.splitext(file)
            if ext.lower() in [".pdf", ".epub", ".mobi"]:
                files_list.append({
                    "relative_path": os.path.relpath(os.path.join(root, file), directory),
                    "filename": file,
                    "fullpath": os.path.join(root, file),
                    "format": ext.lower()
                })
    return files_list

def match_file_to_excel(file_info, excel_books):
    """Busca si el archivo físico ya está registrado en el Excel original de control."""
    fn_title, fn_author = parse_filename(file_info["filename"])
    
    norm_fn_title = clean_string(fn_title)
    norm_fn_author = clean_string(fn_author)
    
    for book in excel_books:
        norm_ex_title = clean_string(book["titulo"])
        norm_ex_author = clean_string(book["autor"])
        
        if norm_fn_title == norm_ex_title and (norm_fn_author == norm_ex_author or norm_fn_author == "desconocido" or norm_ex_author == "desconocido"):
            return book
            
    for book in excel_books:
        norm_ex_title = clean_string(book["titulo"])
        if len(norm_ex_title) > 3 and norm_ex_title in clean_string(file_info["filename"]):
            return book
            
    return None

def main():
    global google_api_enabled
    print("=========================================================")
    print("Iniciando Módulo de Extracción de Metadatos (Biblioteca)")
    print("=========================================================")
    
    env = load_env()
    google_key = env.get("GOOGLE_BOOKS_API_KEY")
    if google_key:
        print(f"[ENV] Clave de API de Google Books detectada en .env.")
    else:
        print(f"[ENV] No se encontró GOOGLE_BOOKS_API_KEY en .env. Se usará el modo anónimo.")
        
    api_cache = load_cache()
    print(f"[CACHE] Se cargaron {len(api_cache)} registros de caché persistente.")
    
    excel_books = load_excel_data(excel_input)
    print(f"[EXCEL] Se cargaron {len(excel_books)} registros del Excel de control original.")
    
    physical_files = scan_physical_files(books_dir)
    print(f"[DISK] Se encontraron {len(physical_files)} archivos de libros en '{books_dir}'.")
    
    processed_catalog = []
    cache_dirty = False
    
    print("\n--- Procesando catálogo de libros ---")
    
    for idx, file_info in enumerate(physical_files, 1):
        rel_path = file_info["relative_path"]
        ext = file_info["format"]
        
        matched_book = match_file_to_excel(file_info, excel_books)
        
        # Extraer metadatos locales antes de definir el título/autor final y la clave de caché
        local_title, local_author, local_subject, local_desc = None, None, None, None
        if ext == ".epub":
            local_title, local_author, local_subject, local_desc = get_epub_metadata(file_info["fullpath"])
        elif ext == ".pdf":
            local_title, local_author, local_subject, local_desc = get_pdf_metadata(file_info["fullpath"])
            
        fn_title, fn_author = parse_filename(file_info["filename"])
        
        final_title = matched_book["titulo"] if matched_book else local_title or fn_title
        final_author = matched_book["autor"] if matched_book else local_author or fn_author
        
        if not final_title: final_title = fn_title
        if not final_author: final_author = fn_author
        final_title = final_title.strip()
        final_author = final_author.strip()
        
        cache_key = clean_string(final_title) + "||" + clean_string(final_author)
        
        suggested_genre = "No especificado"
        synopsis = ""
        source_used = "ninguno"
        
        if cache_key in api_cache:
            c_entry = api_cache[cache_key]
            if c_entry.get("sinopsis") or (c_entry.get("genero") and c_entry.get("genero") != "No especificado") or c_entry.get("source") != "ninguno":
                suggested_genre = c_entry["genero"]
                synopsis = c_entry["sinopsis"]
                source_used = c_entry.get("source", "cache")
                print(f"[{idx}/{len(physical_files)}] [CACHÉ] {final_title} | {final_author}")
                
                original_genre = "No especificado"
                genero_columna_final = suggested_genre
                genero_sugerido_columna = ""
                
                if matched_book:
                    original_genre = matched_book["genero"]
                    final_title = matched_book["titulo"]
                    final_author = matched_book["autor"]
                    genero_columna_final = original_genre
                    if clean_string(suggested_genre) != clean_string(original_genre) and suggested_genre != "No especificado":
                        genero_sugerido_columna = suggested_genre
                else:
                    genero_columna_final = suggested_genre
                    genero_sugerido_columna = suggested_genre
                
                # Evaluar si la metadata está completa (Listo)
                tiene_titulo = bool(final_title and final_title.strip() and final_title.lower() != "desconocido")
                tiene_autor = bool(final_author and final_author.strip() and final_author.lower() != "desconocido")
                tiene_genero = bool(genero_columna_final and genero_columna_final.strip() and genero_columna_final.lower() != "no especificado")
                tiene_sinopsis = bool(synopsis and len(synopsis.strip()) > 15)
                no_funca = "no funca" in rel_path.lower()
                listo = tiene_titulo and tiene_autor and tiene_genero and tiene_sinopsis and not no_funca
                    
                processed_catalog.append({
                    "titulo": final_title,
                    "autor": final_author,
                    "formato": ext,
                    "genero": genero_columna_final,
                    "genero_sugerido": genero_sugerido_columna,
                    "sinopsis": synopsis,
                    "nombre_archivo": rel_path,
                    "listo": listo
                })
                continue
                
        # C. Si no está en caché válida, intentar usar la extracción LOCAL
        if local_desc and len(local_desc.strip()) > 35:
            suggested_genre = map_raw_genre(local_subject) if local_subject else "No especificado"
            synopsis = local_desc.strip()
            source_used = "local_metadata"
            print(f"[{idx}/{len(physical_files)}] [LOCAL] {final_title} | {final_author}")
            print("  -> Extraído localmente desde metadato interno de archivo.")
        else:
            print(f"[{idx}/{len(physical_files)}] [EXTRACT] {final_title} | {final_author}")
            # SÓLO hacer búsquedas si el título NO es puramente numérico y tiene longitud > 2
            is_valid_for_query = True
            clean_title_digits = re.sub(r'[\s_\-\.]', '', final_title)
            if clean_title_digits.isdigit() or len(final_title.strip()) <= 2:
                is_valid_for_query = False
                
            if is_valid_for_query:
                g_genre, g_syn = query_google_books(final_title, final_author, google_key)
                if g_genre or g_syn:
                    suggested_genre = g_genre or "No especificado"
                    synopsis = g_syn or ""
                    source_used = "google_books"
                    print(f"  -> Obtenido de Google Books (Longitud Sinopsis: {len(synopsis)})")
                    time.sleep(1.5)
                    
                if (not suggested_genre or suggested_genre == "No especificado" or not synopsis):
                    w_genre, w_syn = query_wikidata_wikipedia(final_title, final_author)
                    if w_genre or w_syn:
                        suggested_genre = w_genre or suggested_genre or "No especificado"
                        synopsis = w_syn or synopsis or ""
                        source_used = "wikidata_wikipedia"
                        print(f"  -> Obtenido de Wikidata/Wikipedia (Longitud Sinopsis: {len(synopsis)})")
                        time.sleep(1.5)
                        
                if (not suggested_genre or suggested_genre == "No especificado" or not synopsis):
                    ol_genre, ol_syn = query_open_library(final_title, final_author)
                    if ol_genre or ol_syn:
                        suggested_genre = ol_genre or suggested_genre or "No especificado"
                        synopsis = ol_syn or synopsis or ""
                        source_used = "open_library"
                        print(f"  -> Obtenido de Open Library (Longitud Sinopsis: {len(synopsis)})")
                        time.sleep(1.5)
            else:
                print("  -> Omitiendo búsquedas online por título puramente numérico o muy corto.")
                
            if not suggested_genre or suggested_genre == "No especificado":
                suggested_genre = map_raw_genre(local_subject) if local_subject else "No especificado"
            if not synopsis:
                synopsis = local_desc or ""
            if source_used == "ninguno" and (suggested_genre != "No especificado" or synopsis):
                source_used = "local_metadata_fallback"
                
        api_cache[cache_key] = {
            "titulo": final_title,
            "autor": final_author,
            "genero": suggested_genre,
            "sinopsis": synopsis,
            "source": source_used
        }
        cache_dirty = True
        
        if idx % 10 == 0:
            save_cache(api_cache)
            cache_dirty = False
            
        original_genre = "No especificado"
        genero_columna_final = suggested_genre
        genero_sugerido_columna = ""
        
        if matched_book:
            original_genre = matched_book["genero"]
            final_title = matched_book["titulo"]
            final_author = matched_book["autor"]
            genero_columna_final = original_genre
            
            if clean_string(suggested_genre) != clean_string(original_genre) and suggested_genre != "No especificado":
                genero_sugerido_columna = suggested_genre
        else:
            genero_columna_final = suggested_genre
            genero_sugerido_columna = suggested_genre
            
        # Evaluar si la metadata está completa (Listo)
        tiene_titulo = bool(final_title and final_title.strip() and final_title.lower() != "desconocido")
        tiene_autor = bool(final_author and final_author.strip() and final_author.lower() != "desconocido")
        tiene_genero = bool(genero_columna_final and genero_columna_final.strip() and genero_columna_final.lower() != "no especificado")
        tiene_sinopsis = bool(synopsis and len(synopsis.strip()) > 15)
        no_funca = "no funca" in rel_path.lower()
        listo = tiene_titulo and tiene_autor and tiene_genero and tiene_sinopsis and not no_funca
            
        processed_catalog.append({
            "titulo": final_title,
            "autor": final_author,
            "formato": ext,
            "genero": genero_columna_final,
            "genero_sugerido": genero_sugerido_columna,
            "sinopsis": synopsis,
            "nombre_archivo": rel_path,
            "listo": listo
        })
        
    if cache_dirty:
        save_cache(api_cache)
        print("[CACHE] Guardado final de la caché completado.")
        
    print("\nGenerando nuevo archivo Excel consolidado...")
    wb_out = openpyxl.Workbook()
    sheet_out = wb_out.active
    sheet_out.title = "Catálogo Biblioteca"
    
    headers = ["item", "listo", "titulo", "autor", "formato", "genero", "genero_sugerido", "sinopsis", "nombre_archivo"]
    sheet_out.append(headers)
    
    # Estilos para celdas de openpyxl
    from openpyxl.styles import PatternFill, Alignment, Font
    fill_true = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid") # Verde suave
    fill_false = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid") # Amarillo suave
    align_center = Alignment(horizontal="center")
    
    for item_idx, book in enumerate(processed_catalog, 1):
        sheet_out.append([
            item_idx,
            book["listo"],
            book["titulo"],
            book["autor"],
            book["formato"],
            book["genero"],
            book["genero_sugerido"],
            book["sinopsis"],
            book["nombre_archivo"]
        ])
        
        # Aplicar formato a la fila recién añadida (1-based en openpyxl, +1 por cabecera)
        row_num = item_idx + 1
        cell_listo = sheet_out.cell(row=row_num, column=2)
        cell_listo.alignment = align_center
        if book["listo"]:
            cell_listo.fill = fill_true
        else:
            cell_listo.fill = fill_false
            
    for col in sheet_out.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        sheet_out.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)
        
    wb_out.save(excel_output)
    print(f"\n¡Éxito! Catálogo consolidado guardado en: {excel_output}")
    print(f"Total de libros cargados y listos: {len(processed_catalog)}")

if __name__ == "__main__":
    main()
